from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from src.exports import dataframe_to_csv_bytes, dataframe_to_xlsx_bytes


def test_dataframe_to_csv_bytes_includes_utf8_bom_and_rows():
    df = pd.DataFrame([{"title": "Emergency Management Specialist", "score": 91}])

    data = dataframe_to_csv_bytes(df)

    assert data.startswith(b"\xef\xbb\xbf")
    assert b"Emergency Management Specialist" in data


def test_dataframe_to_xlsx_bytes_creates_readable_workbook():
    df = pd.DataFrame(
        [
            {"title": "Emergency Management Specialist", "score": 91},
            {"title": "Program Analyst", "score": 84},
        ]
    )

    data = dataframe_to_xlsx_bytes(df, sheet_name="Scorecards", title="GovJobs Scorecards")
    wb = load_workbook(BytesIO(data), read_only=False, data_only=True)
    ws = wb["Scorecards"]

    assert ws["A1"].value == "GovJobs Scorecards"
    assert ws["A3"].value == "title"
    assert ws["A4"].value == "Emergency Management Specialist"
    assert ws["B5"].value == 84
    assert ws.freeze_panes == "A4"
